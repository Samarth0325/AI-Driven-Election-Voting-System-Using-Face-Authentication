# app.py
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import uuid
import json
import re
import os
import io
from flask import Flask, request, jsonify, session
from datetime import datetime, timedelta
import base64
import numpy as np
import cv2
import face_recognition
import pickle
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import requests
from twilio.rest import Client
import pandas as pd

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///voters.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/face_data'
app.config['ENCODINGS_FOLDER'] = 'face_encodings'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['BACKUP_FOLDER'] = 'backups'

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Create necessary directories
for folder in [app.config['UPLOAD_FOLDER'], app.config['ENCODINGS_FOLDER'], app.config['BACKUP_FOLDER']]:
    os.makedirs(folder, exist_ok=True)

class SMSConfig:
    # Twilio Configuration - REPLACE WITH YOUR ACTUAL CREDENTIALS
    TWILIO_ACCOUNT_SID = 'ACf593b35d3453ef33de896a77170e73d1'
    TWILIO_AUTH_TOKEN = 'f1fa3defcd75033bd5c63fe34f32fda5'
    TWILIO_PHONE_NUMBER = '+12093539505'  # Your Twilio phone number
    
# SMS Service Class
class SMSService:
    def __init__(self):
        self.config = SMSConfig()
        self.sms_enabled = True  # Set to False to disable SMS temporarily
    
    def send_sms_twilio(self, to_number, message):
        """Send SMS using Twilio"""
        try:
            # Check if Twilio credentials are configured
            if (self.config.TWILIO_ACCOUNT_SID.startswith('your_') or 
                self.config.TWILIO_AUTH_TOKEN.startswith('your_')):
                print("Twilio credentials not configured")
                return False
                
            client = Client(self.config.TWILIO_ACCOUNT_SID, self.config.TWILIO_AUTH_TOKEN)
            
            message = client.messages.create(
                body=message,
                from_=self.config.TWILIO_PHONE_NUMBER,
                to=to_number
            )
            
            print(f"SMS sent via Twilio: {message.sid}")
            # Log the SMS activity
            self.log_sms_activity(to_number, 'VERIFICATION', 'SUCCESS', 'Twilio')
            return True
            
        except Exception as e:
            print(f"Twilio SMS error: {str(e)}")
            self.log_sms_activity(to_number, 'VERIFICATION', 'FAILED', f'Twilio: {str(e)}')
            return False
    
    def send_sms_api(self, to_number, message):
        """Send SMS using generic SMS API"""
        try:
            # Check if API credentials are configured
            if self.config.SMS_API_KEY.startswith('your_'):
                print("SMS API credentials not configured")
                return False
                
            payload = {
                'apikey': self.config.SMS_API_KEY,
                'numbers': to_number,
                'message': message,
                'sender': self.config.SMS_SENDER_ID
            }
            
            response = requests.post(self.config.SMS_API_URL, data=payload)
            result = response.json()
            
            if result.get('status') == 'success':
                print(f"SMS sent via API: {result}")
                self.log_sms_activity(to_number, 'VERIFICATION', 'SUCCESS', 'SMS_API')
                return True
            else:
                print(f"SMS API error: {result}")
                self.log_sms_activity(to_number, 'VERIFICATION', 'FAILED', f'SMS_API: {result}')
                return False
                
        except Exception as e:
            print(f"SMS API error: {str(e)}")
            self.log_sms_activity(to_number, 'VERIFICATION', 'FAILED', f'SMS_API: {str(e)}')
            return False
    
    def send_verification_sms(self, voter):
        """Send verification confirmation SMS to voter"""
        if not self.sms_enabled:
            print("SMS service is disabled")
            return False
            
        try:
            message = f"""Hello {voter.name},

Your voter registration has been VERIFIED successfully!

EQIC ID: {voter.eqic_id}
You can now participate in elections using face authentication.

Thank you,
SecureVote Team
"""
            
            # Clean and validate phone number
            phone_number = self.clean_phone_number(voter.phone)
            if not phone_number:
                print("Invalid phone number format")
                return False
            
            print(f"Attempting to send SMS to: {phone_number}")
            
            # Try Twilio first, fallback to API
            success = self.send_sms_twilio(phone_number, message)
            if not success:
                print("Twilio failed, trying SMS API...")
                success = self.send_sms_api(phone_number, message)
            
            return success
            
        except Exception as e:
            print(f"Error sending verification SMS: {str(e)}")
            return False
    
    def clean_phone_number(self, phone):
        """Clean and validate phone number"""
        if not phone:
            return None
            
        # Remove all non-digit characters except +
        cleaned = ''.join(c for c in phone if c.isdigit() or c == '+')
        
        # If no country code, assume +91 (India) - adjust as needed
        if not cleaned.startswith('+'):
            if len(cleaned) == 10:  # Indian mobile number
                cleaned = '+91' + cleaned
            else:
                print(f"Invalid phone number length: {cleaned}")
                return None
        
        return cleaned
    
    def log_sms_activity(self, phone, sms_type, status, provider):
        """Log SMS activity for monitoring"""
        log_entry = f"SMS {sms_type} | To: {phone} | Status: {status} | Provider: {provider}"
        print(f"SMS_LOG: {log_entry}")
        
        # You can also save to database if needed
        # sms_log = SMSLog(phone=phone, type=sms_type, status=status, provider=provider)
        # db.session.add(sms_log)
        # db.session.commit()

# Initialize SMS service
sms_service = SMSService()

# Add SMS Log model (optional - for tracking SMS history)
class SMSLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    voter_id = db.Column(db.Integer, db.ForeignKey('voter.id'), nullable=True)
    phone = db.Column(db.String(20), nullable=False)
    message_type = db.Column(db.String(50), nullable=False)  # VERIFICATION, CUSTOM, etc.
    status = db.Column(db.String(20), nullable=False)  # SUCCESS, FAILED
    provider = db.Column(db.String(50), nullable=True)
    error_message = db.Column(db.Text, nullable=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    
    voter = db.relationship('Voter', backref=db.backref('sms_logs', lazy=True))

class VoteChatbot:
    def __init__(self):
        self.responses = {
            'greeting': {
                'patterns': ['hello', 'hi', 'hey', 'good morning', 'good afternoon'],
                'responses': [
                    "Hello! I'm your SecureVote assistant. How can I help you today?",
                    "Hi there! Ready to assist with your voting queries.",
                    "Welcome to SecureVote! How can I help you?"
                ]
            },
            'voting_process': {
                'patterns': ['how to vote', 'voting process', 'how does voting work', 'voting steps'],
                'responses': [
                    "The voting process is simple:\n1. Register/Login with your credentials\n2. Verify your identity using face authentication\n3. Select your preferred candidate\n4. Review and submit your vote\n5. Receive confirmation",
                    "Here's how to vote:\n- Complete face authentication\n- Choose your candidate\n- Submit your vote securely\n- Get instant confirmation"
                ]
            },
            'face_authentication': {
                'patterns': ['face auth', 'biometric', 'face recognition', 'authentication problem', 'face scan'],
                'responses': [
                    "Face authentication ensures secure voting:\n- Ensure good lighting\n- Look directly at camera\n- Remove accessories if needed\n- If issues persist, contact admin",
                    "For face authentication:\n- Use a well-lit area\n- Position face in frame\n- Hold still during scan\n- Contact support if problems continue"
                ]
            },
            'election_info': {
                'patterns': ['election date', 'when to vote', 'voting time', 'deadline'],
                'responses': [
                    "Election dates are set by administrators. Check the dashboard for current election schedules and deadlines.",
                    "Voting periods vary by election. Please check the active elections section on your dashboard for specific dates."
                ]
            },
            'technical_issues': {
                'patterns': ['not working', 'error', 'bug', 'technical problem', 'page not loading'],
                'responses': [
                    "I'm sorry you're experiencing issues. Try:\n1. Refresh the page\n2. Clear browser cache\n3. Check internet connection\n4. Try different browser\nIf problems persist, contact technical support.",
                    "For technical issues:\n- Refresh your browser\n- Ensure stable internet\n- Try alternate browser\nContact admin if unresolved."
                ]
            },
            'results': {
                'patterns': ['when results', 'result date', 'see results', 'election outcome'],
                'responses': [
                    "Election results are typically announced after the voting period ends. Check the results section for official announcements.",
                    "Results will be available once the election concludes. Administrators will notify voters when results are published."
                ]
            },
            'security': {
                'patterns': ['is it secure', 'security measures', 'vote safety', 'privacy'],
                'responses': [
                    "SecureVote uses multiple security layers:\n- Face biometric authentication\n- Encrypted vote storage\n- Tamper-proof audit logs\n- Real-time monitoring\nYour vote is completely secure and anonymous.",
                    "Security features include:\n- Biometric verification\n- End-to-end encryption\n- Anonymous voting\n- Comprehensive audit trails"
                ]
            },
            'registration': {
                'patterns': ['how to register', 'voter registration', 'sign up', 'create account'],
                'responses': [
                    "Voter registration is handled by administrators. Please contact your election administrator to get registered in the system.",
                    "Registration requires admin approval. Reach out to your election coordinator to be added as a verified voter."
                ]
            },
            'fallback': {
                'patterns': [],
                'responses': [
                    "I'm not sure I understand. Could you rephrase your question?",
                    "I'm still learning! Try asking about voting process, face authentication, or technical help.",
                    "I don't have an answer for that yet. Please contact admin support for specific queries."
                ]
            }
        }
    
    def get_response(self, user_input):
        user_input = user_input.lower().strip()
        
        # Check for specific patterns
        for intent, data in self.responses.items():
            if intent == 'fallback':
                continue
                
            for pattern in data['patterns']:
                if re.search(r'\b' + re.escape(pattern) + r'\b', user_input):
                    return {
                        'response': np.random.choice(data['responses']),
                        'intent': intent,
                        'confidence': 0.9
                    }
        
        # Fallback response
        return {
            'response': np.random.choice(self.responses['fallback']['responses']),
            'intent': 'fallback',
            'confidence': 0.1
        }

# Add this to your Flask app
chatbot = VoteChatbot()

@app.route('/chatbot/message', methods=['POST'])
def chatbot_message():
    try:
        data = request.get_json()
        user_message = data.get('message', '')
        
        if not user_message:
            return jsonify({'error': 'No message provided'}), 400
        
        response = chatbot.get_response(user_message)
        
        # Log the interaction
        log_chatbot_interaction(
            user_id=session.get('voter_id'),
            user_message=user_message,
            bot_response=response['response'],
            intent=response['intent']
        )
        
        return jsonify(response)
    
    except Exception as e:
        return jsonify({'error': 'Chatbot unavailable'}), 500

def log_chatbot_interaction(user_id, user_message, bot_response, intent):
    # Implement logging to database
    pass

# Database Models
class Voter(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    eqic_id = db.Column(db.String(100), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    address = db.Column(db.String(200), nullable=False)
    age = db.Column(db.Integer, nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    face_data_path = db.Column(db.String(200))
    face_encoding_path = db.Column(db.String(200))
    registration_date = db.Column(db.DateTime, default=datetime.utcnow)
    is_verified = db.Column(db.Boolean, default=False)
    has_voted = db.Column(db.Boolean, default=False)
    last_login_attempt = db.Column(db.DateTime)
    login_attempts = db.Column(db.Integer, default=0)
    successful_logins = db.Column(db.Integer, default=0)
    
    def __repr__(self):
        return f'<Voter {self.name}>'

class LoginSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    voter_id = db.Column(db.Integer, db.ForeignKey('voter.id'), nullable=False)
    session_id = db.Column(db.String(100), nullable=False)
    login_time = db.Column(db.DateTime, default=datetime.utcnow)
    success = db.Column(db.Boolean, default=False)
    confidence = db.Column(db.Float, default=0.0)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.Text)
    
    voter = db.relationship('Voter', backref=db.backref('logins', lazy=True))

class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_superadmin = db.Column(db.Boolean, default=False)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Candidate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    party = db.Column(db.String(100), nullable=False)
    
    def __repr__(self):
        return f'<Candidate {self.name}>'

class Vote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    voter_id = db.Column(db.Integer, db.ForeignKey('voter.id'), nullable=False)
    candidate_id = db.Column(db.Integer, db.ForeignKey('candidate.id'), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    
    voter = db.relationship('Voter', backref=db.backref('votes', lazy=True))
    candidate = db.relationship('Candidate', backref=db.backref('votes', lazy=True))

# Add to your models.py
class ChatbotInteraction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    voter_id = db.Column(db.Integer, db.ForeignKey('voter.id'), nullable=True)
    user_message = db.Column(db.Text, nullable=False)
    bot_response = db.Column(db.Text, nullable=False)
    intent = db.Column(db.String(50), nullable=False)
    confidence = db.Column(db.Float, default=0.0)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    resolved = db.Column(db.Boolean, default=False)
    feedback = db.Column(db.Integer, nullable=True)  # 1-5 rating

class ChatbotIntent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    patterns = db.Column(db.JSON, nullable=False)  # Store as JSON array
    responses = db.Column(db.JSON, nullable=False) # Store as JSON array
    usage_count = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)

# Decorators
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def voter_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'voter_id' not in session:
            flash('Please authenticate to access this page.', 'warning')
            return redirect(url_for('authenticate'))
        return f(*args, **kwargs)
    return decorated_function

# Utility Functions
def generate_face_encoding(image_path, eqic_id):
    """Generate face encoding from an image and save it"""
    try:
        image = face_recognition.load_image_file(image_path)
        face_locations = face_recognition.face_locations(image)
        
        if len(face_locations) == 0:
            return None, "No face detected in the image"
        
        face_encodings = face_recognition.face_encodings(image, face_locations)
        
        if len(face_encodings) == 0:
            return None, "Could not generate face encoding"
        
        encoding_path = os.path.join(app.config['ENCODINGS_FOLDER'], f"{eqic_id}.pkl")
        
        with open(encoding_path, 'wb') as f:
            pickle.dump(face_encodings[0], f)
        
        return encoding_path, None
    except Exception as e:
        return None, f"Error generating face encoding: {str(e)}"

def compare_faces(known_encoding_path, unknown_image):
    """Compare a face encoding with a new image"""
    try:
        with open(known_encoding_path, 'rb') as f:
            known_encoding = pickle.load(f)
        
        rgb_unknown_image = cv2.cvtColor(unknown_image, cv2.COLOR_BGR2RGB)
        unknown_face_locations = face_recognition.face_locations(rgb_unknown_image)
        
        if len(unknown_face_locations) == 0:
            return False, 0.0, "No face detected in the image"
        
        unknown_face_encodings = face_recognition.face_encodings(rgb_unknown_image, unknown_face_locations)
        
        if len(unknown_face_encodings) == 0:
            return False, 0.0, "Could not generate face encoding for the unknown image"
        
        results = face_recognition.compare_faces([known_encoding], unknown_face_encodings[0])
        face_distances = face_recognition.face_distance([known_encoding], unknown_face_encodings[0])
        
        confidence = 1 - face_distances[0] if len(face_distances) > 0 else 0
        
        is_match = results[0] if len(results) > 0 else False
        if is_match and confidence > 0.6:
            return True, confidence, "Face matched successfully"
        else:
            return False, confidence, "Face did not match"
            
    except Exception as e:
        return False, 0.0, f"Error comparing faces: {str(e)}"

def generate_voting_excel():
    """Generate Excel file with voting details"""
    try:
        # Query voting data with voter and candidate information
        votes_data = db.session.query(
            Voter.eqic_id,
            Voter.name.label('voter_name'),
            Voter.email,
            Voter.phone,
            Candidate.name.label('candidate_name'),
            Candidate.party,
            Vote.timestamp
        ).join(Vote, Voter.id == Vote.voter_id)\
         .join(Candidate, Candidate.id == Vote.candidate_id)\
         .order_by(Vote.timestamp.desc()).all()

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Voting Details"

        # Headers
        headers = ['EQIC ID', 'Voter Name', 'Email', 'Phone', 'Candidate Name', 'Party', 'Vote Timestamp']
        
        # Style headers
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment

        # Add data
        for row, vote in enumerate(votes_data, 2):
            ws.cell(row=row, column=1, value=vote.eqic_id)
            ws.cell(row=row, column=2, value=vote.voter_name)
            ws.cell(row=row, column=3, value=vote.email)
            ws.cell(row=row, column=4, value=vote.phone)
            ws.cell(row=row, column=5, value=vote.candidate_name)
            ws.cell(row=row, column=6, value=vote.party)
            ws.cell(row=row, column=7, value=vote.timestamp.strftime('%Y-%m-%d %H:%M:%S'))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        return wb

    except Exception as e:
        print(f"Error generating Excel: {e}")
        return None

def auto_backup_votes():
    """Automatically backup voting data"""
    try:
        filename = os.path.join(app.config['BACKUP_FOLDER'], f"votes_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        wb = generate_voting_excel()
        if wb:
            wb.save(filename)
            print(f"Backup created: {filename}")
            return True
        return False
    except Exception as e:
        print(f"Backup error: {e}")
        return False

# Main Application Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        try:
            # Check if it's an AJAX request
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            name = request.form['name']
            address = request.form['address']
            age = int(request.form['age'])
            email = request.form['email']
            phone = request.form['phone']
            face_data = request.form['face_data']
            
            eqic_id = str(uuid.uuid4())[:12].upper()
            
            face_path = None
            encoding_path = None
            if face_data and face_data.startswith('data:image'):
                face_data = face_data.split(',')[1]
                img_data = base64.b64decode(face_data)
                
                face_filename = f"{eqic_id}_face.jpg"
                face_path = os.path.join(app.config['UPLOAD_FOLDER'], face_filename)
                
                with open(face_path, 'wb') as f:
                    f.write(img_data)
                
                encoding_path, encoding_error = generate_face_encoding(face_path, eqic_id)
                if encoding_error:
                    if is_ajax:
                        return jsonify({'success': False, 'message': f'Face encoding error: {encoding_error}'})
                    else:
                        flash(f'Face encoding error: {encoding_error}', 'warning')
            
            new_voter = Voter(
                eqic_id=eqic_id,
                name=name,
                address=address,
                age=age,
                email=email,
                phone=phone,
                face_data_path=face_path,
                face_encoding_path=encoding_path
            )
            
            db.session.add(new_voter)
            db.session.commit()
            
            # Return JSON for AJAX requests, redirect for normal form submissions
            if is_ajax:
                return jsonify({
                    'success': True, 
                    'message': 'Registration successful!',
                    'eqic_id': eqic_id
                })
            else:
                flash(f'Registration successful! Your EQIC ID is: {eqic_id}', 'success')
                return redirect(url_for('registration_success', eqic_id=eqic_id))
            
        except Exception as e:
            db.session.rollback()
            print(f"Error: {e}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': False, 
                    'message': 'Error occurred during registration. Please try again.'
                })
            else:
                flash('Error occurred during registration. Please try again.', 'danger')
    
    return render_template('register.html')

@app.route('/registration_success/<eqic_id>')
def registration_success(eqic_id):
    voter = Voter.query.filter_by(eqic_id=eqic_id).first()
    return render_template('registration_success.html', voter=voter)

@app.route('/authenticate', methods=['GET', 'POST'])
def authenticate():
    if request.method == 'POST':
        eqic_id = request.form.get('eqic_id', '').strip().upper()
        face_data = request.form.get('face_data', '')
        
        if not eqic_id:
            return jsonify({'success': False, 'message': 'Please enter your EQIC ID'})
        
        voter = Voter.query.filter_by(eqic_id=eqic_id).first()
        
        if not voter:
            return jsonify({'success': False, 'message': 'Invalid EQIC ID'})
        
        voter.login_attempts += 1
        voter.last_login_attempt = datetime.utcnow()
        
        if not voter.face_encoding_path or not os.path.exists(voter.face_encoding_path):
            db.session.commit()
            return jsonify({'success': False, 'message': 'No face data available for this voter'})
        
        if not voter.is_verified:
            db.session.commit()
            return jsonify({'success': False, 'message': 'Voter not verified. Please contact administrator.'})
        
        if face_data and face_data.startswith('data:image'):
            try:
                face_data = face_data.split(',')[1]
                img_data = base64.b64decode(face_data)
                
                nparr = np.frombuffer(img_data, np.uint8)
                image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                
                is_match, confidence, message = compare_faces(voter.face_encoding_path, image)
                
                session_id = str(uuid.uuid4())
                login_session = LoginSession(
                    voter_id=voter.id,
                    session_id=session_id,
                    success=is_match,
                    confidence=confidence,
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent')
                )
                db.session.add(login_session)
                
                if is_match:
                    voter.successful_logins += 1
                    
                    session['voter_id'] = voter.id
                    session['voter_name'] = voter.name
                    session['voter_eqic_id'] = voter.eqic_id
                    session['auth_session_id'] = session_id
                    
                    db.session.commit()
                    
                    return jsonify({
                        'success': True, 
                        'message': f'Authentication successful! Welcome {voter.name}',
                        'confidence': round(confidence, 2)
                    })
                else:
                    db.session.commit()
                    return jsonify({
                        'success': False, 
                        'message': f'Authentication failed: {message}',
                        'confidence': round(confidence, 2)
                    })
                    
            except Exception as e:
                db.session.commit()
                return jsonify({'success': False, 'message': f'Error processing image: {str(e)}'})
        else:
            db.session.commit()
            return jsonify({'success': False, 'message': 'Please capture your face image'})
    
    return render_template('authenticate.html')

@app.route('/authentication_success')
def authentication_success():
    if 'voter_id' not in session:
        flash('Please authenticate first.', 'warning')
        return redirect(url_for('authenticate'))
    
    voter = Voter.query.get(session['voter_id'])
    login_session = LoginSession.query.filter_by(session_id=session.get('auth_session_id')).first()
    
    return render_template('authentication_success.html', voter=voter, login_session=login_session)

@app.route('/check_voting_status')
def check_voting_status():
    if 'voter_id' in session:
        voter = Voter.query.get(session['voter_id'])
        return jsonify({
            'can_vote': not voter.has_voted,
            'has_voted': voter.has_voted,
            'is_verified': voter.is_verified
        })
    return jsonify({'can_vote': False, 'has_voted': False, 'is_verified': False})

@app.route('/vote')
@voter_required
def vote():
    candidates = Candidate.query.all()
    return render_template('vote.html', candidates=candidates)

@app.route('/submit_vote/<int:candidate_id>', methods=['POST'])
@voter_required
def submit_vote(candidate_id):
    voter_id = session['voter_id']
    voter = Voter.query.get(voter_id)

    if not voter:
        flash('Voter not found.', 'danger')
        return redirect(url_for('index'))

    if voter.has_voted:
        flash('You have already cast your vote.', 'warning')
        return redirect(url_for('index'))
    
    try:
        vote = Vote(voter_id=voter.id, candidate_id=candidate_id)
        db.session.add(vote)
        
        voter.has_voted = True
        
        db.session.commit()
        
        # Create backup after successful vote
        auto_backup_votes()
        
        flash('Your vote has been successfully cast!', 'success')
    except IntegrityError:
        db.session.rollback()
        flash('There was an error recording your vote.', 'danger')
    
    return redirect(url_for('index'))

@app.route('/vote_splash')
@voter_required
def vote_splash():
    voted_candidate_name = session.pop('voted_candidate_name', None)
    if not voted_candidate_name:
        return redirect(url_for('index'))
    return render_template('vote_splash.html', voted_candidate_name=voted_candidate_name)

@app.route('/results')
def results():
    candidates = Candidate.query.all()
    votes = db.session.query(
        Vote.candidate_id, func.count(Vote.id).label('vote_count')
    ).group_by(Vote.candidate_id).all()
    
    results_dict = {vote.candidate_id: vote.vote_count for vote in votes}
    
    final_results = []
    for candidate in candidates:
        final_results.append({
            'name': candidate.name,
            'party': candidate.party,
            'votes': results_dict.get(candidate.id, 0)
        })
        
    final_results.sort(key=lambda x: x['votes'], reverse=True)
    
    return render_template('results.html', results=final_results)

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))

# Admin Routes
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        admin = Admin.query.filter_by(username=username).first()
        
        if admin and admin.check_password(password):
            session['admin_id'] = admin.id
            session['admin_name'] = admin.name
            session['is_superadmin'] = admin.is_superadmin
            flash('Login successful!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid username or password.', 'danger')
    
    return render_template('admin/login.html')

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    total_voters = Voter.query.count()
    verified_voters = Voter.query.filter_by(is_verified=True).count()
    voted_voters = Voter.query.filter_by(has_voted=True).count()
    
    total_logins = LoginSession.query.count()
    successful_logins = LoginSession.query.filter_by(success=True).count()
    failed_logins = total_logins - successful_logins
    success_rate = (successful_logins / total_logins * 100) if total_logins > 0 else 0
    
    recent_voters = Voter.query.order_by(Voter.registration_date.desc()).limit(5).all()
    recent_logins = LoginSession.query.order_by(LoginSession.login_time.desc()).limit(5).all()
    
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    daily_logins = db.session.query(
        func.date(LoginSession.login_time).label('date'),
        func.count(LoginSession.id).label('count')
    ).filter(LoginSession.login_time >= seven_days_ago).group_by(
        func.date(LoginSession.login_time)
    ).all()
    
    login_dates = [login.date for login in daily_logins]
    login_counts = [login.count for login in daily_logins]
    
    total_votes = Vote.query.count()
    results = db.session.query(
        Candidate.name,
        Candidate.party,
        func.count(Vote.id).label('vote_count')
    ).join(Vote).group_by(Candidate.id).order_by(func.count(Vote.id).desc()).all()
    
    return render_template('admin/dashboard.html', 
                           total_voters=total_voters,
                           verified_voters=verified_voters,
                           voted_voters=voted_voters,
                           recent_voters=recent_voters,
                           total_logins=total_logins,
                           successful_logins=successful_logins,
                           failed_logins=failed_logins,
                           success_rate=round(success_rate, 2),
                           recent_logins=recent_logins,
                           login_dates=login_dates,
                           login_counts=login_counts,
                           total_votes=total_votes,
                           results=results)

@app.route('/admin/voters')
@login_required
def admin_voters():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    voters = Voter.query.order_by(Voter.registration_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False)
    
    return render_template('admin/voters.html', voters=voters)

@app.route('/admin/voter/<int:voter_id>')
@login_required
def voter_details(voter_id):
    voter = Voter.query.get_or_404(voter_id)
    return render_template('admin/voter_details.html', voter=voter)

@app.route('/admin/verify_voter/<int:voter_id>')
@login_required
def verify_voter(voter_id):
    voter = Voter.query.get_or_404(voter_id)
    
    if voter.is_verified:
        flash(f'Voter {voter.name} is already verified.', 'info')
        return redirect(url_for('voter_details', voter_id=voter_id))
    
    voter.is_verified = True
    db.session.commit()
    
    # Send SMS notification
    sms_sent = sms_service.send_verification_sms(voter)
    
    if sms_sent:
        flash(f'Voter {voter.name} has been verified and SMS notification sent!', 'success')
    else:
        flash(f'Voter {voter.name} has been verified but SMS notification failed. Please check phone number format.', 'warning')
    
    return redirect(url_for('voter_details', voter_id=voter_id))
# SMS API Routes
@app.route('/admin/send_verification_sms/<int:voter_id>')
@login_required
def send_verification_sms(voter_id):
    """API endpoint to send verification SMS"""
    voter = Voter.query.get_or_404(voter_id)
    
    if not voter.is_verified:
        return jsonify({'success': False, 'message': 'Voter is not verified'})
    
    sms_sent = sms_service.send_verification_sms(voter)
    
    if sms_sent:
        return jsonify({'success': True, 'message': 'Verification SMS sent successfully!'})
    else:
        return jsonify({'success': False, 'message': 'Failed to send SMS. Please check phone number format and SMS configuration.'})

@app.route('/admin/send_custom_sms/<int:voter_id>', methods=['POST'])
@login_required
def send_custom_sms(voter_id):
    """API endpoint to send custom SMS"""
    voter = Voter.query.get_or_404(voter_id)
    data = request.get_json()
    
    if not data or 'message' not in data:
        return jsonify({'success': False, 'message': 'No message provided'})
    
    custom_message = data['message']
    
    try:
        # Clean phone number
        phone_number = sms_service.clean_phone_number(voter.phone)
        if not phone_number:
            return jsonify({'success': False, 'message': 'Invalid phone number format'})
        
        # Send custom SMS
        success = sms_service.send_sms_twilio(phone_number, custom_message)
        if not success:
            success = sms_service.send_sms_api(phone_number, custom_message)
        
        if success:
            return jsonify({'success': True, 'message': 'Custom SMS sent successfully!'})
        else:
            return jsonify({'success': False, 'message': 'Failed to send SMS. Please check SMS configuration.'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/admin/sms_status')
@login_required
def sms_status():
    """Check SMS service status"""
    return jsonify({
        'sms_enabled': sms_service.sms_enabled,
        'twilio_configured': not SMSConfig.TWILIO_ACCOUNT_SID.startswith('your_'),
        'api_configured': not SMSConfig.SMS_API_KEY.startswith('your_')
    })

@app.route('/admin/delete_voter/<int:voter_id>')
@login_required
def delete_voter(voter_id):
    voter = Voter.query.get_or_404(voter_id)

    # Delete face data files if they exist
    if voter.face_data_path and os.path.exists(voter.face_data_path):
        os.remove(voter.face_data_path)
    if voter.face_encoding_path and os.path.exists(voter.face_encoding_path):
        os.remove(voter.face_encoding_path)

    try:
        # Delete all associated login session records
        LoginSession.query.filter_by(voter_id=voter.id).delete()
        
        # Delete the voter record
        db.session.delete(voter)
        
        db.session.commit()
        flash(f'Voter {voter.name} and their data have been deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting voter: {str(e)}', 'danger')
        
    return redirect(url_for('admin_voters'))

@app.route('/admin/search')
@login_required
def search_voters():
    query = request.args.get('q', '')
    
    if query:
        voters = Voter.query.filter(
            (Voter.name.ilike(f'%{query}%')) | 
            (Voter.email.ilike(f'%{query}%')) |
            (Voter.eqic_id.ilike(f'%{query}%'))
        ).order_by(Voter.registration_date.desc()).all()
    else:
        voters = []
    
    return render_template('admin/search.html', voters=voters, query=query)

@app.route('/admin/auth_logs')
@login_required
def auth_logs():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    logs = LoginSession.query.order_by(LoginSession.login_time.desc()).paginate(
        page=page, per_page=per_page, error_out=False)
    
    return render_template('admin/auth_logs.html', logs=logs)

@app.route('/admin/export_votes_excel')
@login_required
def export_votes_excel():
    """Export voting data to Excel"""
    try:
        wb = generate_voting_excel()
        if wb:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"voting_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('Error generating Excel file', 'danger')
            return redirect(url_for('admin_dashboard'))
            
    except Exception as e:
        flash(f'Error exporting Excel: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/export_votes_csv')
@login_required
def export_votes_csv():
    """Export voting data to CSV"""
    try:
        votes_data = db.session.query(
            Voter.eqic_id,
            Voter.name.label('voter_name'),
            Voter.email,
            Voter.phone,
            Candidate.name.label('candidate_name'),
            Candidate.party,
            Vote.timestamp
        ).join(Vote, Voter.id == Vote.voter_id)\
         .join(Candidate, Candidate.id == Vote.candidate_id)\
         .order_by(Vote.timestamp.desc()).all()

        data = [{
            'EQIC_ID': vote.eqic_id,
            'Voter_Name': vote.voter_name,
            'Email': vote.email,
            'Phone': vote.phone,
            'Candidate_Name': vote.candidate_name,
            'Party': vote.party,
            'Vote_Timestamp': vote.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        } for vote in votes_data]

        df = pd.DataFrame(data)
        
        buffer = io.StringIO()
        df.to_csv(buffer, index=False)
        buffer.seek(0)
        
        filename = f"voting_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return send_file(
            io.BytesIO(buffer.getvalue().encode()),
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )
            
    except Exception as e:
        flash(f'Error exporting CSV: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))

# Error Handlers
@app.errorhandler(413)
def too_large(e):
    return "File is too large. Please use a smaller image or reduce quality.", 413

@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    return render_template('errors/500.html'), 500

# Initialize Database
def initialize_database():
    """Initialize database with default data"""
    with app.app_context():
        db.create_all()
        
        # Create default admin if not exists
        if not Admin.query.filter_by(username='admin').first():
            default_admin = Admin(
                username='admin',
                name='System Administrator',
                email='admin@votingsystem.com',
                is_superadmin=True
            )
            default_admin.set_password('admin123')
            db.session.add(default_admin)
            db.session.commit()
            print("Default admin created: admin / admin123")
        
        # Create default candidates if none exist
        if not Candidate.query.first():
            candidates = [
                Candidate(name='Candidate A', party='Independent'),
                Candidate(name='Candidate B', party='Democratic'),
                Candidate(name='Candidate C', party='Republican')
            ]
            db.session.add_all(candidates)
            db.session.commit()
            print("Default candidates created")

if __name__ == '__main__':
    initialize_database()
    app.run(debug=True)